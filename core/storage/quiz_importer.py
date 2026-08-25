"""
Question BankImportModule - SupportMoreTypeFormatImport
Format：JSON、Excel、Word(.docx)、PDF
"""
import json
import os
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional
from pathlib import Path


@dataclass
class Question:
    """QuestionDataStructure"""
    id: str
    subject: str  # Subject
    grade: str    # Grade
    chapter: str  # Chapter
    question_type: str  # QuestionType：Single Choice/MoreSelect/Judge/Fill-in-blank
    question: str  # QuestionInsideContent
    options: List[str] = None  # Option（SelectQuestion）
    answer: str = ""  # Answer
    explanation: str = ""  # Explanation
    source: str = ""  # SourceFile


class QuizImporter:
    """Question BankImport"""
    
    def __init__(self):
        self.questions: List[Question] = []
        self.import_history = []
    
    def import_file(self, filepath: str) -> Dict:
        """ImportSingleItemFile，BackResult"""
        ext = Path(filepath).suffix.lower()
        
        try:
            if ext == '.json':
                return self._import_json(filepath)
            elif ext in ['.xlsx', '.xls']:
                return self._import_excel(filepath)
            elif ext == '.docx':
                return self._import_docx(filepath)
            elif ext == '.pdf':
                return self._import_pdf(filepath)
            else:
                return {"success": False, "error": f"NotSupportFormat: {ext}"}
        except Exception as e:
            import traceback
            error_msg = f"ImportFailed: {str(e)}"
            print(f"[DEBUG] ImportError: {error_msg}")
            print(f"[DEBUG] Stack: {traceback.format_exc()}")
            return {"success": False, "error": error_msg}
    
    def _import_json(self, filepath: str) -> Dict:
        """ImportJSONFormatQuestion Bank"""
        # TryMoreEncoding Type
        encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030', 'latin-1']
        
        data = None
        used_encoding = None
        
        for enc in encodings:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    data = json.load(f)
                    used_encoding = enc
                    break
            except (json.JSONDecodeError, UnicodeDecodeError):
                continue
        
        if data is None:
            return {"success": False, "error": f"CannotExplanationJSONFile，TryEncoding: {', '.join(encodings)}"}
        
        questions = []
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            items = data.get('questions', [data])
        else:
            items = []
        
        for item in items:
            q = self._parse_question(item, source=filepath)
            if q:
                questions.append(q)
        
        self.import_history.append({
            "file": filepath,
            "count": len(questions),
            "encoding": used_encoding,
            "time": self._get_time()
        })
        
        return {"success": True, "count": len(questions), "questions": questions}
    
    def _import_excel(self, filepath: str) -> Dict:
        """ImportExcelFormatQuestion Bank"""
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "error": "Need To Install openpyxl: pip install openpyxl"}
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        questions = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(headers, row))
            q = self._parse_question(item, source=filepath)
            if q:
                questions.append(q)
        
        self.import_history.append({
            "file": filepath,
            "count": len(questions),
            "time": self._get_time()
        })
        
        return {"success": True, "count": len(questions), "questions": questions}
    
    def _import_docx(self, filepath: str) -> Dict:
        """ImportWordFormatQuestion Bank"""
        try:
            from docx import Document
        except ImportError:
            return {"success": False, "error": "Need To Install python-docx: pip install python-docx"}
        
        doc = Document(filepath)
        questions = []
        
        # ByParagraphExplanation
        text = "\n".join([p.text for p in doc.paragraphs])
        q_list = self._extract_questions_from_text(text)
        
        for item in q_list:
            q = self._parse_question(item, source=filepath)
            if q:
                questions.append(q)
        
        self.import_history.append({
            "file": filepath,
            "count": len(questions),
            "time": self._get_time()
        })
        
        return {"success": True, "count": len(questions), "questions": questions}
    
    def _import_pdf(self, filepath: str) -> Dict:
        """ImportPDFFormatQuestion Bank"""
        try:
            import pdfplumber
        except ImportError:
            return {"success": False, "error": "Need To Install pdfplumber: pip install pdfplumber"}
        
        questions = []
        
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                q_list = self._extract_questions_from_text(text)
                for item in q_list:
                    q = self._parse_question(item, source=filepath)
                    if q:
                        questions.append(q)
        
        self.import_history.append({
            "file": filepath,
            "count": len(questions),
            "time": self._get_time()
        })
        
        return {"success": True, "count": len(questions), "questions": questions}
    
    def _extract_questions_from_text(self, text: str) -> List[Dict]:
        """FromTextInExtractQuestion"""
        import re
        
        questions = []
        # MatchQuestionFormat：CountCharacter+QuestionNumber+QuestionInsideContent
        pattern = r'(\d+)[\.、）\)]\s*([^0-9][^AnswerOption]*?)(?=Answer|[（\(]|$)'
        
        for match in re.finditer(pattern, text, re.DOTALL):
            q_num = match.group(1)
            q_text = match.group(2).strip()[:200]
            
            # TryExtractAnswer
            answer_match = re.search(r'Answer[:：]?\s*([A-Dabcd])', q_text)
            answer = answer_match.group(1) if answer_match else ""
            
            questions.append({
                "question": f"{q_num}. {q_text}",
                "answer": answer,
                "subject": "",
                "grade": "",
                "chapter": ""
            })
        
        return questions
    
    def _parse_question(self, item: Dict, source: str = "") -> Optional[Question]:
        """ExplanationQuestionData"""
        try:
            # GetQuestionInsideContent - CompatibleContentMoreTypeFormat
            question_text = ""
            
            # Format1: DirectField
            if 'question' in item:
                question_text = item['question']
            elif 'QuestionDry' in item:
                question_text = item['QuestionDry']
            elif 'title' in item:
                question_text = item['title']
            # Format2: NestedStructure（Such AsEducationalAPI）
            elif 'question_info' in item and isinstance(item['question_info'], dict):
                q_info = item['question_info']
                question_text = q_info.get('raw_content', '') or q_info.get('content', '') or q_info.get('question', '')
            
            if not question_text:
                return None
            
            # GetOption
            options = []
            if 'options' in item:
                options = item['options']
            elif 'Option' in item:
                options = item['Option']
            elif 'answer_options' in item:
                options = item['answer_options']
            if isinstance(options, str):
                options = [options]
            
            # GetAnswer
            answer = ""
            if 'answer' in item:
                answer = item['answer']
            elif 'Answer' in item:
                answer = item['Answer']
            elif 'answer_info' in item and isinstance(item['answer_info'], dict):
                answer = item['answer_info'].get('raw_content', '') or item['answer_info'].get('content', '')
            
            # GetSubject
            subject = item.get('subject', item.get('Subject', ''))
            if not subject and 'course' in item:
                subject = item['course']
            
            # GetGrade
            grade = item.get('grade', item.get('Grade', ''))
            
            # GetQuestionType
            question_type = item.get('question_type', item.get('QuestionType', 'Not yetKnowledge'))
            if not question_type or question_type == 'Not yetKnowledge':
                if 'type' in item:
                    question_type = item['type']
            
            # GetExplanation
            explanation = item.get('explanation', item.get('Explanation', ''))
            if not explanation and 'solution_info' in item:
                sol = item['solution_info']
                if isinstance(sol, list) and len(sol) > 0:
                    if isinstance(sol[0], dict):
                        explanation = sol[0].get('solution_info', '') or sol[0].get('content', '')
                elif isinstance(sol, dict):
                    explanation = sol.get('solution_info', '') or sol.get('content', '')
            
            # GenerateID - PriorityUseJSONInOriginalID
            q_id = item.get('id', f"import_{int(__import__('time').time()*1000)}_{len(self.questions)}")
            
            return Question(
                id=q_id,
                subject=str(subject),
                grade=str(grade),
                chapter=item.get('chapter', item.get('Chapter', '')),
                question_type=str(question_type)[:20],
                question=str(question_text)[:500],
                options=options if isinstance(options, list) else [],
                answer=str(answer)[:100],
                explanation=str(explanation)[:500],
                source=source
            )
        except Exception:
            return None
    
    def _get_time(self):
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M")
    
    def get_all_questions(self) -> List[Dict]:
        """Get AllQuestion（Used ForSave）"""
        return [asdict(q) for q in self.questions]
    
    def save_to_db(self, db_path: str):
        """SaveToDataLibrary"""
        import sqlite3
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        for q in self.questions:
            cursor.execute('''
                INSERT OR REPLACE INTO questions 
                (id, subject, grade, chapter, type, question, options, answer, explanation, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (q.id, q.subject, q.grade, q.chapter, q.question_type,
                  q.question, json.dumps(q.options, ensure_ascii=False),
                  q.answer, q.explanation, q.source))
        
        conn.commit()
        conn.close()
